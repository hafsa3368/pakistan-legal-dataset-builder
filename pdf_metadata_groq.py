import pymupdf
import json
import logging
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os, re, time
import easyocr
from PIL import Image
import numpy as np
import requests

load_dotenv()

# ================= CONFIG =================

MODEL      = "llama3.2:3b"
OLLAMA_URL = "http://localhost:11434/api/generate"

SCRIPT_DIR   = Path(__file__).parent    
ZIP_FILES    = ["pdfs.zip", "shc.zip"]
EXTRACT_DIR  = SCRIPT_DIR / "extracted_pdfs"
OUTPUT_EXCEL = SCRIPT_DIR / "supreme_court_ai_metadata.xlsx"
CHECKPOINT   = SCRIPT_DIR / "checkpoint.json"
MAPPING_EXCEL = SCRIPT_DIR / "pdf_name_metadata.xlsx"   # NEW: actual_filename/generated_name -> actual_path mapping. Naam alag ho to yahan badal do.

DELAY       = 0      # FIX: 1 sec tha — Groq rate-limit ke zamane ka leftover, local Ollama ko zaroorat nahi
MAX_PAGES   = 2
MAX_RETRIES = 1
SAVE_EVERY  = 1      # FIX: har PDF ke baad save (pehle 25 tha) — Ctrl+C pe kuch bhi na khoye

HEADERS = ["File Name", "Actual File Path", "Court", "Case Type", "Case Number", "Year", "Legal Issue", "Keywords", "Summary"]
# NEW: "Actual File Path" — asal PDF ka pura disk path. Ab future mein kabhi "PDF nahi mili" wala
# masla nahi hoga, kyunki path khud Excel mein save hai, dobara dhoondhna nahi padega.
# NEW: "Case Number" — raw case citation (jaise "Cr.B.A.No.S-825 of 2020") "Case Type"
# se alag column mein rehta hai. "Case Type" sirf clean category rakhta hai (Bail,
# Criminal Appeal, waghera).

# FIX: folder (root.name) se actual court pata karne ke liye — pehle hardcoded "COURT" jaata tha
COURT_MAP = {
    "pdfs": "Lahore High Court",
    "shc":  "High Court of Sindh",
}

logging.basicConfig(filename=str(SCRIPT_DIR / "errors.log"), level=logging.ERROR)

# ================= NORMALIZATION HELPERS =================
# NEW: Ye saare functions ab shuru se hi lagte hain — taake naye process hone
# wale PDFs ka data pehli dafa mein hi saaf (consistent) aaye, aur baad mein
# alag se "fix" script chalane ki zaroorat na pade.

def normalize_court(raw, fallback="Unknown"):
    """Court field ko standard naam mein normalize karta hai (Ollama har baar
    thoda different likhta hai — "HIGH COURT OF SINDH", "High Court of Sindh,
    Circuit Court, Hyderabad" waghera, sab ek standard naam mein aate hain).
    Naya court format nazar aaye to neeche ek 'if' badha dena."""
    if not raw or not str(raw).strip():
        return fallback

    s = str(raw).upper()

    if "SINDH" in s:
        return "High Court of Sindh"
    if "LAHORE" in s:
        return "Lahore High Court"
    if "SUPREME COURT" in s:
        return "Supreme Court of Pakistan"
    if "BALOCHISTAN" in s:
        return "Balochistan High Court"
    if "PESHAWAR" in s:
        return "Peshawar High Court"
    if "ISLAMABAD" in s:
        return "Islamabad High Court"

    # Koi known court match nahi hui. Ab check karo ye asal mein court ka naam
    # lagta hai ya OCR garbage ("Ord-r vith Sig-4*ure c/ Judga" jaisa).
    has_court_word = "COURT" in s
    junk_chars = sum(1 for ch in s if ch in "-*/_~^|\\")
    looks_garbled = junk_chars >= 2

    if not has_court_word or looks_garbled:
        return fallback

    return str(raw).strip()


def normalize_case_type(raw, fallback="Unknown"):
    """Raw case-type text ko chhoti si standard category list mein map karta
    hai. Substring-matching hai is liye "Cr. Appeal", "Crl. Appeal", "Criminal
    Appeal" — sab automatically "Criminal Appeal" ban jate hain. Naya
    abbreviation/pattern nazar aaye to neeche ek 'if' badha dena."""
    if not raw or not str(raw).strip():
        return fallback

    s = str(raw).upper()
    s_compact = re.sub(r"[^A-Z]", "", s)   # dots/numbers/spaces hata ke sirf letters

    if "BAIL" in s:
        return "Bail"
    if re.search(r"\bB\.?A\.?\b", raw) or "CRBA" in s_compact or "MBA" in s_compact:
        return "Bail"   # "Cr.B.A.No...", "Crl.B.A" jaisi abbreviations
    if "APPEAL" in s:
        return "Civil Appeal" if "CIVIL" in s else "Criminal Appeal"
    if "REVISION" in s:
        return "Criminal Revision"
    if "WRIT" in s:
        return "Writ Petition"
    if "CONSTITUTION" in s:
        return "Constitutional Petition"
    if "SUIT" in s:
        return "Civil Suit"
    if "REFERENCE" in s:
        return "Reference"
    if "APPLICATION" in s:
        return "Criminal Application"

    return fallback


def normalize_year(raw, fallback="Unknown"):
    """Kabhi kabhi AI poori date de deta hai ("12.10.2019") jabke baaki sab
    jagah sirf saal hota hai ("2020"). Hamesha sirf 4-digit saal nikaal ke deta hai."""
    if not raw:
        return fallback

    m = re.search(r"(19|20)\d{2}", str(raw))
    if m:
        return m.group()

    return fallback


def normalize_keywords(raw_keywords):
    """Keywords list ko clean karta hai — extra whitespace/dots hatata hai,
    duplicate (case-insensitive) hatata hai, comma-separated string wapas deta hai."""
    if not raw_keywords:
        return ""

    if isinstance(raw_keywords, str):
        items = raw_keywords.split(",")
    elif isinstance(raw_keywords, list):
        items = raw_keywords
    else:
        return ""

    seen = set()
    cleaned = []
    for kw in items:
        kw = str(kw).strip().strip(".").strip()
        if not kw:
            continue
        key = kw.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(kw)

    return ", ".join(cleaned)


# NEW: Pakistani court judgments apna case number aam taur par pehle page pe
# khud likhte hain (jaise "Cr.B.A. No. S-825 of 2020"). Ye regex se seedha PDF
# text se nikalta hai — AI se zyada reliable hai kyunki AI kabhi poora number
# chhod deta hai ("bail" likh ke reh jata hai).
CASE_NUMBER_PATTERNS = [
    r"Cr\.?\s*B\.?A\.?\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"Crl\.?\s*Misc\.?\s*Application\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"Cr\.?\s*Misc\.?\s*Application\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"Cr\.?\s*Appeal\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"Crl\.?\s*Appeal\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"Crl\.?\s*Revision\s*Application\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"Const(?:itutional)?\.?\s*Petition\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"C\.?P\.?\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"Civil\s*Suit\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    # NEW: purana fallback KISI BHI text ko "No." se pehle utha leta tha — is se
    # kabhi kabhi judge ka naam ("Hassan Shah") case number ke saath chipak jata
    # tha ("ssan Shah Criminal Bail Application No.1283 of 2025"). Ab ye
    # WHITELIST-based fallback hai — "No." se pehle SIRF known legal words
    # (Bail, Appeal, Application, Revision, HCA, waghera) allow hain, koi
    # random naam/text nahi aa sakta.
    (
        r"(?:(?:Const(?:itutional)?|Civil|Criminal|Crl|Cr|C\.P\.?|CP|Sessions|Bail|"
        r"Application|Appl|Appeal[s]?|Revision|Petition|Suit|Case|Jail|"
        r"Acq(?:uittal)?|HCA|Reference|Writ|Misc(?:ellaneous)?|Intra-?Court|"
        r"Service|Rev|A\.?T\.?|B\.?A\.?)\.?\s*){1,6}"
        r"No\.?s?\.?\s*[:\-]?\s*[A-Z]?\s*[\-\u2013]?\s*\d+[\-/]?[A-Z]?\d*\s*(?:of|/)\s*\d{4}"
    ),
]
COMPILED_CASE_NUMBER_PATTERNS = [re.compile(p, re.IGNORECASE) for p in CASE_NUMBER_PATTERNS]


def extract_case_number_from_text(text):
    """PDF text se regex ke zariye real case number dhoondta hai. Nahi mila to None."""
    if not text:
        return None
    flat_text = re.sub(r"\s+", " ", text)
    for pattern in COMPILED_CASE_NUMBER_PATTERNS:
        m = pattern.search(flat_text)
        if m:
            return m.group().strip()
    return None


def looks_like_real_case_number(value):
    """Check karta hai ke value asal case number jaisi lagti hai ya generic/garbage
    hai — "No." + digit + 4-digit saal teeno maujood hon to real samjho."""
    if not value:
        return False
    s = str(value)
    has_no    = re.search(r"\bno\.?s?\.?\b", s, re.IGNORECASE) is not None
    has_year  = re.search(r"(19|20)\d{2}", s) is not None
    has_digit = re.search(r"\d", s) is not None
    return has_no and has_year and has_digit

# ================= OCR =================

_OCR = None

def get_ocr():
    global _OCR
    if _OCR is None:
        print("[*] EasyOCR initialize ho raha hai...")
        _OCR = easyocr.Reader(['en'], gpu=False)
    return _OCR

def ocr_image(pixmap):
    try:
        ocr = get_ocr()
        img = Image.frombytes("RGB", (pixmap.width, pixmap.height), pixmap.samples)
        arr = np.array(img)
        res = ocr.readtext(arr)
        return "\n".join([r[1] for r in res if r[1].strip()])
    except Exception as e:
        logging.error(f"OCR error: {str(e)}")
        return ""

# ================= UTILS =================

# FIX: in columns mein se koi bhi khali/Unknown ho to row "incomplete" maani jayegi
def is_incomplete(row):
    if not row.get("Court") or row.get("Court") == "Unknown":
        return True
    return any(not row.get(c) for c in ["Case Type", "Year", "Legal Issue", "Keywords", "Summary"])

# NEW: mapping excel (actual_filename/generated_name -> actual_path) sirf ek baar load karo,
# har row ke liye dobara Excel kholna bohot slow hota
_MAPPING = None
_FILE_INDEX = None   # NEW: filename -> Path index, ek hi baar build hota hai

def build_file_index():
    """EXTRACT_DIR ko sirf EK BAAR poora scan karo aur filename->path dict bana lo.
    Pehle har 'PDF nahi mili' check pe poora folder dobara scan ho raha tha —
    2000 missing files ke liye 2000 baar 60k files scan = bohot slow."""
    global _FILE_INDEX
    if _FILE_INDEX is not None:
        return _FILE_INDEX

    print("[*] PDF index ban raha hai (ek baar hoga, thora time lagega)...")
    _FILE_INDEX = {}
    for p in EXTRACT_DIR.rglob("*.pdf"):
        _FILE_INDEX.setdefault(p.name, p)
    print(f"[i] Index mein {len(_FILE_INDEX)} PDFs mile.")
    return _FILE_INDEX

def load_mapping():
    global _MAPPING
    if _MAPPING is not None:
        return _MAPPING

    _MAPPING = {}
    if not MAPPING_EXCEL.exists():
        print(f"[!] Mapping file nahi mili: {MAPPING_EXCEL} — sirf rglob fallback chalega (slow).")
        return _MAPPING

    wb = load_workbook(MAPPING_EXCEL)
    ws = wb.active
    header = [c.value for c in ws[1]]
    col_idx = {h: i for i, h in enumerate(header)}

    for r in range(2, ws.max_row + 1):
        actual_filename = ws.cell(row=r, column=col_idx["actual_filename"] + 1).value
        actual_path     = ws.cell(row=r, column=col_idx["actual_path"] + 1).value
        generated_name  = None
        if "generated_name" in col_idx:
            generated_name = ws.cell(row=r, column=col_idx["generated_name"] + 1).value

        if actual_path:
            if actual_filename:
                _MAPPING[actual_filename] = actual_path
            if generated_name:
                _MAPPING[generated_name] = actual_path

    print(f"[i] Mapping file se {len(_MAPPING)} filename → path entries load hue.")
    return _MAPPING

# FIX: Excel mein sirf filename hota hai — pehle mapping file se direct path nikalo (fast),
# agar wahan na mile to purana rglob fallback (slow, par PDF mil to jayegi)
def find_pdf(filename):
    mapping = load_mapping()

    real_path = mapping.get(filename)
    if real_path:
        p = Path(real_path)
        if p.exists():
            return p

    index = build_file_index()             # FIX: O(1) lookup, har baar dobara folder scan nahi
    return index.get(filename)

def clean(text):
    return re.sub(r'[^a-zA-Z0-9_-]', '_', str(text)).strip('_')

def safe_json(txt):
    txt = re.sub(r"```(?:json)?", "", txt).strip()

    try:
        return json.loads(txt)
    except:
        pass

    m = re.search(r"\{.*\}", txt, re.DOTALL)
    if m:
        try:
            return json.loads(m.group())
        except:
            return None

    return None

# ================= PDF READ =================

def read_pdf_text(path):
    try:
        doc = pymupdf.open(str(path))
        parts = []

        for i in range(min(MAX_PAGES, len(doc))):
            text = doc[i].get_text("text")

            text = "\n".join([
                l.strip() for l in text.split("\n")
                if len(l.strip()) > 2 and not l.strip().isdigit()
            ])

            if not text or len(text) < 100:
                print(f"    [*] OCR use ho raha hai: page {i+1}")
                pix = doc[i].get_pixmap()
                text = ocr_image(pix)

            parts.append(text)

        doc.close()
        return "\n\n--- PAGE ---\n\n".join(parts)

    except Exception as e:
        logging.error(str(e))
        return ""

# ================= AI (OLLAMA FIXED) =================

def ask_ai(text, filename=None):

    prompt = f"""
Extract legal metadata and return ONLY valid JSON:

{{
"court":"",
"case_type":"",
"year":"",
"legal_issue":"",
"keywords":[],
"summary":""
}}

RULES:
- No explanation
- No markdown
- JSON only
- "legal_issue" aur "summary" hamesha bharo, chahe document chhota ya procedural ho — kam se kam
  ek line ka summary zaroor do (jaise "Bail application under Section 497 CrPC" ya
  "Interim order adjourning the hearing")
- "keywords" mein kam se kam 2-3 relevant legal terms do, khali list mat chhodo

TEXT:
{text[:6000]}
"""

    last_parsed = None

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            res = requests.post(
                OLLAMA_URL,
                json={
                    "model": MODEL,
                    "prompt": prompt,
                    "stream": False
                },
                timeout=120
            )

            res.raise_for_status()
            data = res.json()

            raw = data.get("response", "")
            parsed = safe_json(raw)

            if parsed:
                last_parsed = parsed

                # FIX: pehle valid JSON aate hi "success" maan liya jata tha, chahe
                # legal_issue/keywords/summary teeno khali hon. Ab aisi "weak" extraction
                # pe ek aur koshish hogi (agar retries baqi hain).
                is_weak = not parsed.get("summary")   # FIX: ab sirf Summary khali hone par bhi retry hoga

                # NEW: agar Summary khali hai (chahe baqi fields bhar gaye hon), raw response
                # debug file mein save karo — taake pata chale Ollama asal mein kya bhej raha hai
                if not parsed.get("summary"):
                    try:
                        with open(SCRIPT_DIR / "ai_debug.log", "a", encoding="utf-8") as dbg:
                            dbg.write(f"\n--- {filename} (attempt {attempt}) ---\n{raw}\n")
                    except Exception:
                        pass

                if not is_weak or attempt == MAX_RETRIES:
                    if is_weak:
                        print(f"    [debug] Weak extraction, Ollama ka raw jawab tha:\n    {raw[:500]}")
                    return parsed, False

                logging.error(f"Weak extraction ({filename}) attempt {attempt}, retry kar rahe hain")
                time.sleep(1)
                continue

            logging.error(f"JSON parse failed ({filename}) attempt {attempt}")
            time.sleep(2)

        except Exception as e:
            logging.error(f"AI error ({filename}) attempt {attempt}: {e}")
            msg = str(e).lower()

            if "connection" in msg or "refused" in msg:
                print("\n[!] Ollama band hai → run: ollama serve\n")
                return None, True

            time.sleep(3)

    return last_parsed, False   # FIX: retries khatam hone par bhi jo aakhri parse mila wo lautao (Court/Year to mil chuka hoga)

# ================= PROCESS PDF =================

def process_pdf(path, folder, tp):

    text = read_pdf_text(path)

    year_guess = ""
    m = re.search(r"(19|20)\d{2}", path.name)
    if m:
        year_guess = m.group()

    # NEW: regex se seedha PDF text se real case number nikalne ki koshish karo
    regex_case_number = extract_case_number_from_text(text)

    row = {
        "File Name": path.name,             # FIX: pehle "" tha — kabhi set hi nahi hota tha, Excel mein filename khali aata
        "Actual File Path": str(path.resolve()),   # NEW: pura path save — future mein dhoondhna hi nahi padega
        "Court": normalize_court(tp, fallback=tp),
        "Case Type": normalize_case_type(path.parent.name, fallback=path.parent.name),
        "Case Number": regex_case_number or "",
        "Year": normalize_year(year_guess, fallback=year_guess),
        "Legal Issue": "",
        "Keywords": "",
        "Summary": "",
    }

    ai_ok = False

    if text:
        ai, stop = ask_ai(text, filename=path.name)

        if stop:
            return row, False, True

        if ai:
            row["Court"] = normalize_court(ai.get("court", "") or tp, fallback=tp)

            raw_case_type = ai.get("case_type", "")
            row["Case Type"] = normalize_case_type(raw_case_type, fallback=row["Case Type"])

            # NEW: Case Number ke liye priority — (1) regex se PDF text se mila real
            # number (sabse reliable), (2) agar wo na mile to AI ka raw case_type
            # (agar wo khud real number jaisa lagta ho), (3) warna khali chhod do
            if regex_case_number:
                row["Case Number"] = regex_case_number
            elif looks_like_real_case_number(raw_case_type):
                row["Case Number"] = raw_case_type
            else:
                row["Case Number"] = row["Case Number"] or ""

            row["Year"]        = normalize_year(ai.get("year", "") or year_guess, fallback=year_guess)
            row["Legal Issue"] = ai.get("legal_issue", "")
            row["Keywords"]    = normalize_keywords(ai.get("keywords", []))
            row["Summary"]     = ai.get("summary", "")
            ai_ok = True

    return row, ai_ok, False

# ================= EXCEL =================

def save_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Metadata"

    for i, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=i, value=h).font = Font(bold=True)

    for r in rows:
        ws.append([r.get(h, "") for h in HEADERS])

    wb.save(OUTPUT_EXCEL)
    print(f"[+] Excel saved: {len(rows)} rows")

# NEW: jin rows ki PDF kahin nahi mili (mapping na rglob index mein) — review ke liye alag file
def save_unresolved(rows):
    path = SCRIPT_DIR / "unresolved_rows.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Unresolved"
    for i, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=i, value=h).font = Font(bold=True)
    for r in rows:
        ws.append([r.get(h, "") for h in HEADERS])
    wb.save(path)
    print(f"[+] {len(rows)} unresolved rows yahan save hui: {path}")

# FIX: purani excel se rows wapas load karne ke liye (resume ke waqt data na khoye)
def load_existing_rows():
    if not OUTPUT_EXCEL.exists():
        return []
    wb = load_workbook(OUTPUT_EXCEL)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {h: (ws.cell(row=r, column=i + 1).value or "") for i, h in enumerate(header)}
        rows.append(row)
    return rows

# FIX: checkpoint load/save — pehle CHECKPOINT variable defined tha par kabhi use nahi hota tha
def load_checkpoint():
    if CHECKPOINT.exists():
        try:
            with open(CHECKPOINT, "r", encoding="utf-8") as f:
                return set(json.load(f).get("processed", []))
        except Exception:
            return set()
    return set()

def save_checkpoint(processed_set):
    with open(CHECKPOINT, "w", encoding="utf-8") as f:
        json.dump({"processed": list(processed_set)}, f)

# ================= MAIN =================

def run():
    all_rows  = load_existing_rows()                       # FIX: purana data load karo, overwrite na ho
    processed = load_checkpoint()

    if not processed and all_rows:
        processed = {r.get("File Name") for r in all_rows if r.get("File Name")}

    # FIX: pehle existing rows mein se incomplete walon ko fill karo (Year/Court/Legal Issue/Keywords/Summary khali)
    incomplete = [r for r in all_rows if is_incomplete(r)]
    print(f"[i] Pehle se saved rows mein incomplete: {len(incomplete)}")

    # FIX: pehle total/remaining count kahin nahi pata chalta tha — ab upfront count kar lo
    pdf_list = []
    for z in ZIP_FILES:
        root = EXTRACT_DIR / Path(z).stem
        if root.exists():
            pdf_list.append((root, list(root.rglob("*.pdf"))))

    total_pdfs = sum(len(p) for _, p in pdf_list)
    print(f"[i] Total PDFs mile: {total_pdfs}")
    print(f"[i] Pehle se ho chuki: {len(processed)}")

    todo_count = sum(1 for _, pdfs in pdf_list for pdf in pdfs if pdf.name not in processed)
    print(f"[i] Abhi process karni hain: {todo_count}\n")

    n = 0
    unresolved = []   # NEW: jo rows ki PDF kahin nahi mili (purana corrupt filename) — review ke liye alag rakho
    try:                                              # FIX: try/finally — Ctrl+C ya kisi bhi crash pe bhi save zaroor ho

        # ===== PHASE 1: purani incomplete rows fill karo =====
        for idx, row in enumerate(incomplete, 1):
            filename = row.get("File Name")
            if not filename:
                continue

            remaining_fill = len(incomplete) - idx
            print(f"[fill {idx}/{len(incomplete)}] {filename}  (remaining: {remaining_fill})")

            # NEW: pehle se saved values bhi normalize kar do — chahe AI dobara na bhi chale
            # to bhi ye row Excel mein standard format ke saath save hogi
            if row.get("Court"):
                row["Court"] = normalize_court(row["Court"], fallback="Unknown")

            if row.get("Case Type") and not row.get("Case Number"):
                # Raw text (agar case-number-jaisa lage) "Case Number" mein rakh lo
                row["Case Number"] = row["Case Type"]
                row["Case Type"]   = normalize_case_type(row["Case Type"], fallback=row["Case Type"])

            if row.get("Year"):
                row["Year"] = normalize_year(row["Year"], fallback=row["Year"])

            pdf_path = None
            saved_path = row.get("Actual File Path")          # NEW: agar pura path pehle se save hai, seedha wahi use karo
            if saved_path and Path(saved_path).exists():
                pdf_path = Path(saved_path)
            else:
                pdf_path = find_pdf(filename)                  # purani (mapping/index) rows ke liye fallback

            if not pdf_path:
                print("    [!] PDF nahi mili, skip")
                unresolved.append(row)
                continue

            row["Actual File Path"] = str(pdf_path.resolve())  # NEW: future ke liye ab yahan se hamesha mil jayega

            text = read_pdf_text(pdf_path)
            if not text:
                print("    [!] Text/OCR dono fail, skip")
                continue

            # NEW: agar Case Number abhi bhi generic/garbage lage, regex se
            # dobara real number nikalne ki koshish karo
            if not looks_like_real_case_number(row.get("Case Number")):
                regex_case_number = extract_case_number_from_text(text)
                if regex_case_number:
                    row["Case Number"] = regex_case_number

            ai, stop = ask_ai(text, filename=filename)

            if stop:
                save_excel(all_rows)
                save_checkpoint(processed)
                print("[!] Ollama band ho gaya — 'ollama serve' chala ke yehi script dobara chalao.")
                return

            if ai:
                if not row.get("Court") or row.get("Court") == "Unknown":
                    row["Court"] = normalize_court(ai.get("court", ""), fallback="Unknown")
                if not row.get("Case Type") or row.get("Case Type") == "Unknown":
                    raw_case_type = ai.get("case_type", "")
                    if not looks_like_real_case_number(row.get("Case Number")) and looks_like_real_case_number(raw_case_type):
                        row["Case Number"] = raw_case_type
                    row["Case Type"] = normalize_case_type(raw_case_type, fallback="Unknown")
                if not row.get("Year"):
                    row["Year"] = normalize_year(ai.get("year", ""), fallback="Unknown")
                if not row.get("Legal Issue"):
                    row["Legal Issue"] = ai.get("legal_issue", "")
                if not row.get("Keywords"):
                    row["Keywords"] = normalize_keywords(ai.get("keywords", []))
                if not row.get("Summary"):
                    row["Summary"] = ai.get("summary", "")
                print("    [+] Filled")
            else:
                print("    [!] AI fail hui")

            if idx % SAVE_EVERY == 0:
                save_excel(all_rows)
                save_checkpoint(processed)

        if incomplete:
            save_excel(all_rows)
            print(f"[+] Incomplete rows wala phase khatam.")
            print(f"[i] PDF na milne wali rows: {len(unresolved)} — ye purane buggy run ka leftover lag rahi hain.")
            if unresolved:
                save_unresolved(unresolved)
            print()

        # ===== PHASE 2: naye PDFs process karo =====
        for root, pdfs in pdf_list:
            tp = COURT_MAP.get(root.name, "Unknown")   # FIX: hardcoded "COURT" ki jagah folder se actual court

            for pdf in pdfs:

                if pdf.name in processed:              # FIX: resume — already done PDFs skip
                    continue

                n += 1
                remaining = todo_count - n
                print(f"[{n}/{todo_count}] Processing: {pdf.name}  (remaining: {remaining})")

                try:                                    # FIX: ek bad PDF pura run na gira de
                    row, ok, stop = process_pdf(pdf, root.name, tp)
                except KeyboardInterrupt:
                    raise                               # upar wale finally tak jaane do, yahan catch nahi karna
                except Exception as e:
                    logging.error(f"Unexpected error on {pdf.name}: {e}")
                    print(f"    [!] Unexpected error, skip kar diya: {e}")
                    continue

                if stop:
                    save_excel(all_rows)
                    save_checkpoint(processed)
                    print("[!] Ollama band ho gaya — 'ollama serve' chala ke yehi script dobara chalao, wahi se resume hoga.")
                    return

                all_rows.append(row)
                processed.add(pdf.name)

                if ok:
                    print("[+] OK")
                else:
                    print("[!] AI failed")

                if n % SAVE_EVERY == 0:                  # FIX: periodic save — crash hone pe sara kaam na jaye
                    save_excel(all_rows)
                    save_checkpoint(processed)

                time.sleep(DELAY)

    except KeyboardInterrupt:
        print("\n[!] Ctrl+C dabaya gaya — ruk raha hoon, ab tak ka kaam save kar raha hoon...")
    finally:
        save_excel(all_rows)          # FIX: chahe Ctrl+C ho, crash ho, ya normal khatam ho — yahan se save zaroor hoga
        save_checkpoint(processed)
        if unresolved:
            save_unresolved(unresolved)
        print(f"[+] {len(processed)} PDFs ka kaam save ho gaya. Dobara chalao to yahin se resume hoga.")

# ================= ENTRY =================

if __name__ == "__main__":
    EXTRACT_DIR.mkdir(exist_ok=True)
    run()