import os
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────
# CONFIG — apna path yahan set karo
# ─────────────────────────────────────────────
LHC_FOLDER   = r"D:\hafsa_thesis material\supreme_court_scraper\extracted_pdfs\pdfs"
SHC_FOLDER   = r"D:\hafsa_thesis material\supreme_court_scraper\extracted_pdfs\shc"
OUTPUT_EXCEL = r"D:\hafsa_thesis material\supreme_court_scraper\pdf_name_metadata.xlsx"

# ─────────────────────────────────────────────
# GENERATED NAME LOGIC (same as existing pipeline)
# ─────────────────────────────────────────────
def extract_year(filename):
    """
    FIXED: the old version used re.search() which returns the FIRST
    4-digit year-like number in the filename. That broke on filenames
    like "bail_general_lhc_1908_2023LHC2950.pdf", where "1908" is an
    Act/Ordinance reference (Explosive Substances Act, 1908) baked into
    the case-type folder name, and the REAL case year (2023) sits later,
    right before the court citation code "LHC2950".

    Strategy:
      1) Look for a year immediately followed by a court-citation code
         (uppercase letters + a digit), e.g. "2023LHC2950" -> 2023.
         This pattern is unambiguous wherever it exists.
      2) If that pattern isn't present, fall back to the FIRST 4-digit
         year in the filename -- this preserves correct behaviour for
         the other naming convention used in this dataset, e.g.
         "2020_SHC_HYD_1961.pdf" -> 2020 (where 1961 is just a serial
         number, not a year, and there's no citation-code suffix to
         disambiguate it with).
    """
    m = re.search(r'((?:19|20)\d{2})(?=[A-Z]{2,6}\d)', filename)
    if m:
        return m.group(1)

    m = re.search(r'(19|20)\d{2}', filename)
    return m.group() if m else "unknown"

def detect_court(filename, folder_court):
    """LHC ya SHC detect karo filename ya folder se"""
    fn = filename.lower()
    if "lhc" in fn:
        return "LHC"
    if "shc" in fn or "hyderabad" in fn or "karachi" in fn or "hyd" in fn or "khi" in fn:
        return "SHC"
    return folder_court

def detect_case_type(filename, subfolder):
    """Case type detect karo"""
    fn = filename.lower()
    known_types = ["bail", "civil", "criminal", "service", "tax", "writ", "contempt"]

    # Pehle filename se try karo
    for t in known_types:
        if fn.startswith(t + "_") or f"_{t}_" in fn:
            return t

    # Phir subfolder se
    if subfolder.lower() in known_types:
        return subfolder.lower()

    return subfolder.lower() or "unknown"

def make_generated_name(actual_filename, court, case_type, year):
    """
    Existing pipeline jaisa generated name banao:
    e.g. LHC_civil_2024_civil_general_lhc_20_case_1.pdf
    """
    stem = Path(actual_filename).stem
    # Special chars clean karo
    safe_stem = re.sub(r'[^a-zA-Z0-9_\-]', '_', stem)[:60]
    parts = [court, case_type]
    if year != "unknown":
        parts.append(year)
    parts.append(safe_stem)
    return "_".join(parts) + ".pdf"


# ─────────────────────────────────────────────
# SCAN FOLDER
# ─────────────────────────────────────────────
def scan_folder(folder_path, folder_court):
    """Recursively scan folder aur har PDF ka metadata nikalo"""
    records = []
    base    = Path(folder_path)

    if not base.exists():
        print(f"⚠  Folder nahi mila: {folder_path}")
        return records

    for root, dirs, files in os.walk(folder_path):
        for file in sorted(files):
            if not file.lower().endswith(".pdf"):
                continue

            full_path = os.path.join(root, file)

            # Subfolder (category) nikalo
            rel      = os.path.relpath(root, folder_path)
            parts    = rel.split(os.sep)
            subfolder = next(
                (p for p in parts if p.lower() not in (".", "pdfs", "shc", "lhc")),
                "root"
            )

            year       = extract_year(file)
            court      = detect_court(file, folder_court)
            case_type  = detect_case_type(file, subfolder)
            gen_name   = make_generated_name(file, court, case_type, year)

            records.append({
                "actual_filename" : file,
                "actual_path"     : full_path,
                "subfolder"       : subfolder,
                "court"           : court,
                "case_type"       : case_type,
                "year"            : year,
                "generated_name"  : gen_name,
            })

    return records


# ─────────────────────────────────────────────
# SAVE TO EXCEL
# ─────────────────────────────────────────────
def save_to_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Name Metadata"

    headers = [
        "actual_filename",
        "actual_path",
        "subfolder",
        "court",
        "case_type",
        "year",
        "generated_name",
    ]

    # Header row styling
    header_fill = PatternFill("solid", fgColor="2E4057")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_fill and header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    # Data rows
    row_font = Font(name="Arial", size=10)
    for record in records:
        ws.append([record.get(h, "") for h in headers])

    # Apply font to data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = row_font

    # Column widths
    widths = [50, 80, 15, 10, 15, 10, 65]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(records) + 1}"

    wb.save(OUTPUT_EXCEL)
    print(f"\n✅ Excel saved: {OUTPUT_EXCEL}")
    print(f"   Total rows : {len(records)}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def run():
    print("=" * 55)
    print("  PDF Name Metadata Generator")
    print("=" * 55)

    all_records = []

    print(f"\n📂 Scanning LHC folder...")
    lhc = scan_folder(LHC_FOLDER, "LHC")
    print(f"   ✔ Found: {len(lhc)} PDFs")
    all_records.extend(lhc)

    print(f"\n📂 Scanning SHC folder...")
    shc = scan_folder(SHC_FOLDER, "SHC")
    print(f"   ✔ Found: {len(shc)} PDFs")
    all_records.extend(shc)

    print(f"\n📊 Total PDFs: {len(all_records)}")

    # Category breakdown
    from collections import Counter
    cats = Counter(r["subfolder"] for r in all_records)
    print("\n📁 By subfolder:")
    for cat, cnt in sorted(cats.items()):
        print(f"   {cat:15} : {cnt}")

    courts = Counter(r["court"] for r in all_records)
    print("\n⚖  By court:")
    for court, cnt in sorted(courts.items()):
        print(f"   {court:10} : {cnt}")

    save_to_excel(all_records)
    print(f"\n✔  Done! Excel file kholo:")
    print(f"   {OUTPUT_EXCEL}")


if __name__ == "__main__":
    run()