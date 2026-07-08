"""
fix_existing_metadata.py
=========================
Ye script tumhari MOJOODA supreme_court_ai_metadata.xlsx ko EK HI RUN mein
poori tarah clean karti hai — do alag steps hain, dono isi file mein, sahi
order mein chalte hain:

STEP 1 — normalize (koi PDF nahi khulti, sirf Excel ke andar):
  1. Court column normalize (HIGH COURT OF SINDH, High Court of Sindh,
     IN THE HIGH COURT OF SINDH... sab ek standard naam mein)
  2. Case Type column se raw case-number text ko "Case Number" column mein
     move karti hai, Case Type mein sirf clean category rakhti hai
  3. Year column normalize (poori date ho to sirf 4-digit saal)

STEP 2 — case number backfill (PDFs khulti hain, koi AI call nahi):
  Jin rows mein "Case Number" abhi bhi generic/garbage hai (jaise "bail",
  "Criminal Misc.", "ORDER SHEET"), unki PDF ka pehla page khol ke regex se
  real case number dhoondti hai.

Koi row delete nahi hoti. Ye script SEEDHA supreme_court_ai_metadata.xlsx
ko edit karti hai — koi backup ya alag copy nahi banati, koi naya excel
file nahi banati.

Chalane se pehle:
1. Excel file BAND kar do (agar khuli hai to save fail ho jayega)
2. Ye script usi folder mein rakho jahan supreme_court_ai_metadata.xlsx hai
"""

import re
from pathlib import Path
from openpyxl import load_workbook
import pymupdf

# ================= CONFIG =================

SCRIPT_DIR   = Path(__file__).parent
OUTPUT_EXCEL = SCRIPT_DIR / "supreme_court_ai_metadata.xlsx"
SAVE_EVERY   = 100   # STEP 2 mein har 100 rows ke baad progress print

# ================= NORMALIZE FUNCTIONS =================
# (pdf_metadata_groq.py se hoobahoo copy — future mein wahan koi naya
# pattern add ho to yahan bhi add kar dena)

def normalize_court(raw, fallback="Unknown"):
    if not raw or not str(raw).strip():
        return fallback

    s = str(raw).upper()

    if "SINDH" in s:
        return "High Court of Sindh"
    if "LAHORE" in s:
        return "Lahore High Court"
    if "SUPREME COURT" in s:
        return "Supreme Court of Pakistan"
    if "BALOCHISTAN" in s:
        return "Balochistan High Court"
    if "PESHAWAR" in s:
        return "Peshawar High Court"
    if "ISLAMABAD" in s:
        return "Islamabad High Court"

    has_court_word = "COURT" in s
    junk_chars = sum(1 for ch in s if ch in "-*/_~^|\\")
    looks_garbled = junk_chars >= 2

    if not has_court_word or looks_garbled:
        return fallback

    return str(raw).strip()


def normalize_case_type(raw, fallback="Unknown"):
    if not raw or not str(raw).strip():
        return fallback

    s = str(raw).upper()
    s_compact = re.sub(r"[^A-Z]", "", s)

    if "BAIL" in s:
        return "Bail"
    if re.search(r"\bB\.?A\.?\b", raw) or "CRBA" in s_compact or "MBA" in s_compact:
        return "Bail"
    if "APPEAL" in s:
        return "Civil Appeal" if "CIVIL" in s else "Criminal Appeal"
    if "REVISION" in s:
        return "Criminal Revision"
    if "WRIT" in s:
        return "Writ Petition"
    if "CONSTITUTION" in s:
        return "Constitutional Petition"
    if "SUIT" in s:
        return "Civil Suit"
    if "REFERENCE" in s:
        return "Reference"
    if "APPLICATION" in s:
        return "Criminal Application"

    return fallback


def normalize_year(raw, fallback="Unknown"):
    if not raw:
        return fallback

    m = re.search(r"(19|20)\d{2}", str(raw))
    if m:
        return m.group()

    return fallback


def looks_like_real_case_number(value):
    """Check karta hai ke value asal case number jaisi lagti hai ya
    generic/garbage hai - 'No.' + digit + 4-digit saal teeno maujood hon
    to real samjho, warna generic (chahe wo 'bail', 'Criminal Misc.',
    'ORDER SHEET' kuch bhi ho)."""
    if not value:
        return False
    s = str(value)
    has_no    = re.search(r"\bno\.?s?\.?\b", s, re.IGNORECASE) is not None
    has_year  = re.search(r"(19|20)\d{2}", s) is not None
    has_digit = re.search(r"\d", s) is not None
    return has_no and has_year and has_digit


# ================= CASE NUMBER REGEX (STEP 2 ke liye) =================

CASE_NUMBER_PATTERNS = [
    r"Cr\.?\s*B\.?A\.?\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"Crl\.?\s*Misc\.?\s*Application\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"Cr\.?\s*Misc\.?\s*Application\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"Cr\.?\s*Appeal\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"Crl\.?\s*Appeal\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"Crl\.?\s*Revision\s*Application\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"Const(?:itutional)?\.?\s*Petition\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"C\.?P\.?\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    r"Civil\s*Suit\s*No\.?\s*[:\-]?\s*[A-Z]?-?\d+[\-/]?\d*\s+of\s+\d{4}",
    # NEW: purana fallback KISI BHI text ko "No." se pehle utha leta tha — is se
    # kabhi kabhi judge ka naam ("Hassan Shah") case number ke saath chipak jata
    # tha ("ssan Shah Criminal Bail Application No.1283 of 2025"). Ab ye
    # WHITELIST-based fallback hai — "No." se pehle SIRF known legal words
    # (Bail, Appeal, Application, Revision, HCA, waghera) allow hain, koi
    # random naam/text nahi aa sakta.
    (
        r"(?:(?:Const(?:itutional)?|Civil|Criminal|Crl|Cr|C\.P\.?|CP|Sessions|Bail|"
        r"Application|Appl|Appeal[s]?|Revision|Petition|Suit|Case|Jail|"
        r"Acq(?:uittal)?|HCA|Reference|Writ|Misc(?:ellaneous)?|Intra-?Court|"
        r"Service|Rev|A\.?T\.?|B\.?A\.?)\.?\s*){1,6}"
        r"No\.?s?\.?\s*[:\-]?\s*[A-Z]?\s*[\-\u2013]?\s*\d+[\-/]?[A-Z]?\d*\s*(?:of|/)\s*\d{4}"
    ),
]
COMPILED_PATTERNS = [re.compile(p, re.IGNORECASE) for p in CASE_NUMBER_PATTERNS]


def extract_case_number(pdf_path, max_pages=2):
    """PDF ke pehle 'max_pages' pages se case number dhoondhta hai (kabhi
    number pehle page par nahi, doosre page par hota hai). Nahi mila to None."""
    try:
        doc = pymupdf.open(str(pdf_path))
        if len(doc) == 0:
            doc.close()
            return None
        pages_text = []
        for i in range(min(max_pages, len(doc))):
            pages_text.append(doc[i].get_text("text"))
        text = "\n".join(pages_text)
        doc.close()
    except Exception:
        return None

    flat_text = re.sub(r"\s+", " ", text)

    for pattern in COMPILED_PATTERNS:
        m = pattern.search(flat_text)
        if m:
            return m.group().strip()

    return None


# ================= STEP 1: Court / Case Type / Year normalize =================

def normalize_columns(ws):
    header = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(header)}

    if "Court" not in col_idx or "Case Type" not in col_idx or "Year" not in col_idx:
        print("[!] Zaroori columns nahi mile. Header check karo:", header)
        return col_idx

    if "Case Number" not in col_idx:
        insert_at = col_idx["Case Type"] + 1
        ws.insert_cols(insert_at)
        ws.cell(row=1, column=insert_at, value="Case Number")
        header = [c.value for c in ws[1]]
        col_idx = {h: i + 1 for i, h in enumerate(header)}
        print("[+] 'Case Number' column add ho gaya.")

    court_col       = col_idx["Court"]
    case_type_col   = col_idx["Case Type"]
    case_number_col = col_idx["Case Number"]
    year_col        = col_idx["Year"]

    court_changed = 0
    case_changed  = 0
    year_changed  = 0
    unmatched_courts = set()

    for r in range(2, ws.max_row + 1):
        court_cell = ws.cell(row=r, column=court_col)
        old_court = court_cell.value
        new_court = normalize_court(old_court, fallback="Unknown")
        if new_court != old_court:
            court_cell.value = new_court
            court_changed += 1
        if new_court not in {"High Court of Sindh", "Lahore High Court",
                              "Supreme Court of Pakistan", "Balochistan High Court",
                              "Peshawar High Court", "Islamabad High Court", "Unknown"}:
            unmatched_courts.add(new_court)

        case_type_cell   = ws.cell(row=r, column=case_type_col)
        case_number_cell = ws.cell(row=r, column=case_number_col)

        old_case_type = case_type_cell.value
        existing_case_number = case_number_cell.value

        raw_source = existing_case_number or old_case_type

        if raw_source:
            if not existing_case_number:
                case_number_cell.value = old_case_type

            new_case_type = normalize_case_type(raw_source, fallback=raw_source)
            if new_case_type != old_case_type:
                case_type_cell.value = new_case_type
                case_changed += 1

        year_cell = ws.cell(row=r, column=year_col)
        old_year = year_cell.value
        new_year = normalize_year(old_year, fallback=old_year)
        if new_year != old_year:
            year_cell.value = new_year
            year_changed += 1

    print(f"[+] STEP 1 -- Court: {court_changed} rows update hui")
    print(f"[+] STEP 1 -- Case Type/Case Number split: {case_changed} rows mein category clean hui")
    print(f"[+] STEP 1 -- Year: {year_changed} rows update hui")

    if unmatched_courts:
        print(f"\n[i] {len(unmatched_courts)} Court values kisi known pattern se match nahi hui (manually check kar lena):")
        for u in sorted(unmatched_courts, key=str):
            print("   -", u)

    return col_idx


# ================= STEP 2: PDF se real case number backfill =================

def fill_case_numbers(ws, col_idx):
    required = ["Case Number", "Actual File Path"]
    missing = [c for c in required if c not in col_idx]
    if missing:
        print(f"[!] STEP 2 skip -- ye columns nahi mile: {missing}")
        return

    case_number_col = col_idx["Case Number"]
    path_col        = col_idx["Actual File Path"]

    candidates = []
    for r in range(2, ws.max_row + 1):
        current_val = ws.cell(row=r, column=case_number_col).value
        if not looks_like_real_case_number(current_val):
            candidates.append(r)

    print(f"\n[i] STEP 2 -- {len(candidates)} rows mein generic Case Number hai, PDF se dhoondte hain...\n")

    found = 0
    not_found = 0

    for idx, r in enumerate(candidates, 1):
        pdf_path_val = ws.cell(row=r, column=path_col).value
        filename = ws.cell(row=r, column=1).value

        if not pdf_path_val or not Path(pdf_path_val).exists():
            not_found += 1
            continue

        case_no = extract_case_number(Path(pdf_path_val))

        if case_no:
            ws.cell(row=r, column=case_number_col, value=case_no)
            found += 1
            print(f"[{idx}/{len(candidates)}] {filename} -- mil gaya: {case_no}")
        else:
            not_found += 1

        if idx % SAVE_EVERY == 0:
            print(f"    [+] Progress: {idx}/{len(candidates)}")

    print(f"\n[+] STEP 2 mukammal. {found} case numbers mil gaye, {not_found} nahi mile.")


# ================= MAIN =================

def fix_existing_excel():
    if not OUTPUT_EXCEL.exists():
        print(f"[!] Excel file nahi mili: {OUTPUT_EXCEL}")
        return

    wb = load_workbook(OUTPUT_EXCEL)
    ws = wb.active

    col_idx = normalize_columns(ws)
    wb.save(OUTPUT_EXCEL)
    print(f"[+] STEP 1 ke baad file save ho gayi.\n")

    fill_case_numbers(ws, col_idx)
    wb.save(OUTPUT_EXCEL)
    print(f"\n[+] Poora kaam mukammal. File save ho gayi: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    fix_existing_excel()