from openpyxl import load_workbook
from pathlib import Path

OUTPUT_EXCEL = Path("supreme_court_ai_metadata.xlsx")

wb = load_workbook(OUTPUT_EXCEL, data_only=True)
ws = wb.active
headers = [c.value for c in ws[1]]
print("Headers:", headers)
print("Total rows in Excel:", ws.max_row - 1)

# First 3 rows
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True)):
    print(f"Row {i+1}: {row[0]}")  # File Name column

wb.close()